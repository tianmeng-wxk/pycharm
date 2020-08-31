from openpyxl import load_workbook


class ExcelUtil:
    def __init__(self, path):
        self.path = path
        self.excel = load_workbook(path)
        self.sheet = self.excel.active

    def get_sheet_by_index(self, index):
        return self.excel.sheetnames[index]

    def set_sheet_by_index(self, index):
        self.sheet = self.excel.sheetnames[index]

    def get_cell_value(self, rownum, column):
        return self.sheet.cell(rownum, column).value

    def set_cell_value(self, rownum, column, value):
        self.sheet.cell(rownum, column, value=value)

    def get_max_row(self):
        return self.sheet.max_row


if __name__ == '__main__':
    path = '../data/某某模块的测试用例集.xlsx'
    excel = ExcelUtil(path)
    print(excel.get_cell_value(5,5))